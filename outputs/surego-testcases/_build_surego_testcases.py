from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path(__file__).with_name("SureGo体验版完整闭环测试用例.xlsx")


THEME = {
    "navy": "0F172A",
    "blue": "2563EB",
    "cyan": "E0F2FE",
    "green": "16A34A",
    "red": "DC2626",
    "amber": "F59E0B",
    "gray": "64748B",
    "line": "CBD5E1",
    "header": "EEF2FF",
}

thin = Side(style="thin", color=THEME["line"])
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws, table_name=None):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=THEME["navy"])
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=10, color="111827")

    for row_idx in range(2, max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    if table_name and max_row >= 2 and max_col >= 2:
        ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    ws.auto_filter.ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"


def set_widths(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


def add_validations(ws, status_col="I", priority_col="C", start_row=2):
    end_row = max(ws.max_row, start_row)
    dv_status = DataValidation(type="list", formula1='"未测,通过,失败,阻塞,跳过"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_priority)
    dv_status.add(f"{status_col}{start_row}:{status_col}{end_row}")
    dv_priority.add(f"{priority_col}{start_row}:{priority_col}{end_row}")

    for row in range(start_row, end_row + 1):
        cell = f"{status_col}{row}"
        ws.conditional_formatting.add(
            cell,
            FormulaRule(formula=[f'${status_col}{row}="通过"'], fill=PatternFill("solid", fgColor="DCFCE7")),
        )
        ws.conditional_formatting.add(
            cell,
            FormulaRule(formula=[f'${status_col}{row}="失败"'], fill=PatternFill("solid", fgColor="FEE2E2")),
        )
        ws.conditional_formatting.add(
            cell,
            FormulaRule(formula=[f'${status_col}{row}="阻塞"'], fill=PatternFill("solid", fgColor="FEF3C7")),
        )


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


e2e_headers = [
    "用例ID", "模块", "优先级", "测试类型", "前置条件", "测试步骤", "预期结果",
    "实际结果", "状态", "缺陷ID", "负责人", "测试时间", "备注"
]

e2e_rows = [
    ["SG-E2E-001", "体验版启动", "P0", "冒烟", "体验版已上传，测试微信号为体验成员", "扫码进入体验版小程序", "小程序正常打开，首页无白屏、无编译错误", "", "未测", "", "", "", ""],
    ["SG-E2E-002", "微信登录", "P0", "功能", "A 为体验成员，uni-id 已上传并配置", "A 点击底部 + 或报名等登录入口，进入登录页并授权", "登录成功，写入 uni_id_token，返回原目标页", "", "未测", "", "", "", ""],
    ["SG-E2E-003", "登录态持久化", "P1", "回归", "A 已登录", "关闭小程序后重新打开，进入用户中心", "仍识别为已登录，不重复强制授权", "", "未测", "", "", "", ""],
    ["SG-E2E-004", "首页浏览", "P0", "功能", "小程序已启动", "查看首页顶部入口、精选活动、我的局、活动卡片", "内容正常加载，卡片可点击，无明显错位", "", "未测", "", "", "", ""],
    ["SG-E2E-005", "发现页浏览", "P0", "功能", "首页正常", "点击底部右侧发现", "进入发现页，分类、城市、推荐流正常显示", "", "未测", "", "", "", ""],
    ["SG-E2E-006", "搜索入口", "P1", "功能", "首页或发现页可用", "进入搜索页，输入关键词，点击结果", "搜索结果可显示，点击进入对应详情页", "", "未测", "", "", "", ""],
    ["SG-E2E-007", "城市筛选", "P1", "功能", "发现页可用", "点击城市入口，选择城市后返回", "城市选择生效，发现页保留筛选状态", "", "未测", "", "", "", ""],
    ["SG-E2E-008", "日历入口", "P1", "功能", "发现页或用户中心可用", "进入日历页，选择日期活动", "按日期展示活动，活动可进入详情或参与者中心", "", "未测", "", "", "", ""],
    ["SG-E2E-009", "活动详情", "P0", "功能", "存在活动数据", "点击任意活动卡片", "详情页展示封面、标题、时间、地点、人数、局长信息、报名 CTA", "", "未测", "", "", "", ""],
    ["SG-E2E-010", "创建活动", "P0", "主流程", "A 已登录", "底部 + 进入创建，填写标题、分类、时间、地点、人数、费用、介绍、封面", "表单可编辑，必填校验合理，无页面报错", "", "未测", "", "", "", ""],
    ["SG-E2E-011", "位置选择", "P1", "能力", "A 在创建页", "点击选择位置，允许或拒绝定位/位置能力", "可选择位置；失败时有提示，文本地点仍可用", "", "未测", "", "", "", ""],
    ["SG-E2E-012", "封面上传", "P1", "能力", "A 在创建页", "选择图片作为活动封面", "图片上传至云存储或失败提示清晰，不把临时路径作为最终业务数据", "", "未测", "", "", "", ""],
    ["SG-E2E-013", "活动预览", "P1", "功能", "创建表单已填写", "点击预览", "预览内容与活动详情信息一致或高度接近", "", "未测", "", "", "", ""],
    ["SG-E2E-014", "发布活动", "P0", "主流程", "创建表单合法", "点击发布/创建", "跳转成功页，活动写入云端，CTA 可进入管理或详情", "", "未测", "", "", "", ""],
    ["SG-E2E-015", "创建后可见", "P0", "数据一致性", "A 已发布活动", "回到首页、发现、我的活动查看", "新活动可在相关列表中看到，状态一致", "", "未测", "", "", "", ""],
    ["SG-E2E-016", "B 登录", "P0", "功能", "B 为体验成员", "B 扫体验版进入并登录", "B 登录成功，uid 与 A 不同", "", "未测", "", "", "", ""],
    ["SG-E2E-017", "B 浏览 A 活动", "P0", "主流程", "A 已创建活动", "B 从首页/发现进入 A 活动详情", "B 可看到详情，CTA 为报名/申请", "", "未测", "", "", "", ""],
    ["SG-E2E-018", "B 提交报名", "P0", "主流程", "B 在活动详情页", "点击报名，填写申请信息/问题答案并提交", "报名申请写入云端，进入成功页或待审核状态", "", "未测", "", "", "", ""],
    ["SG-E2E-019", "重复报名拦截", "P1", "异常", "B 已提交报名", "B 再次进入详情点击报名", "不重复创建申请，提示已报名/审核中/已加入", "", "未测", "", "", "", ""],
    ["SG-E2E-020", "A 查看报名", "P0", "主流程", "B 已报名", "A 进入我的活动或活动管理", "管理页可看到 B 的申请，统计正确", "", "未测", "", "", "", ""],
    ["SG-E2E-021", "A 审核通过", "P0", "主流程", "管理页有 B 申请", "A 点击通过，填写备注确认", "B 申请状态变为 approved，管理页人数/状态更新", "", "未测", "", "", "", ""],
    ["SG-E2E-022", "审核消息", "P1", "通知", "A 已审核通过", "B 打开消息页", "B 收到审核通过/活动状态消息", "", "未测", "", "", "", ""],
    ["SG-E2E-023", "参与者中心", "P0", "主流程", "B 已通过审核", "B 从详情或我的活动进入参与者中心", "显示活动凭证、报名状态、订单状态、签到入口", "", "未测", "", "", "", ""],
    ["SG-E2E-024", "模拟支付入口", "P0", "主流程", "活动为诚意金/门票类型", "B 点击支付/确认订单", "进入支付页，显示金额、类型、无真实扣款说明", "", "未测", "", "", "", ""],
    ["SG-E2E-025", "模拟支付成功", "P0", "主流程", "B 在支付页", "点击确认支付/试运营确认", "订单状态变为 paid，返回参与者中心", "", "未测", "", "", "", ""],
    ["SG-E2E-026", "订单一致性", "P0", "数据一致性", "B 已支付", "查看参与者中心、用户中心订单 Tab、订单详情", "同一订单状态均为已支付，金额和活动信息一致", "", "未测", "", "", "", ""],
    ["SG-E2E-027", "订单详情", "P1", "功能", "B 有订单", "点击订单卡片进入详情", "展示订单号、状态时间线、金额、活动信息、规则说明", "", "未测", "", "", "", ""],
    ["SG-E2E-028", "A 票券统计", "P1", "管理", "B 已支付", "A 进入管理页票券/保证金统计", "显示待支付、已支付、退款/关闭数量和金额统计", "", "未测", "", "", "", ""],
    ["SG-E2E-029", "签到页进入", "P0", "主流程", "A 为活动创建者", "A 在管理页点击签到核销", "进入签到页，展示活动信息、统计、核销码、参与者列表", "", "未测", "", "", "", ""],
    ["SG-E2E-030", "B 入场凭证", "P0", "主流程", "B 已支付/已加入", "B 打开参与者中心", "展示入场凭证、核销码或签到状态", "", "未测", "", "", "", ""],
    ["SG-E2E-031", "A 手动核销", "P0", "主流程", "B 未签到", "A 输入合法核销码或选择待核销成员确认", "B 签到成功，已签到人数 +1", "", "未测", "", "", "", ""],
    ["SG-E2E-032", "A 扫码核销", "P1", "能力", "存在未签到参与者", "A 点击扫码/模拟扫码", "真机调起扫码；失败或模拟器下有 fallback；核销状态更新", "", "未测", "", "", "", ""],
    ["SG-E2E-033", "重复签到拦截", "P1", "异常", "B 已签到", "A 再次核销 B，或 B 再次点击签到", "不新增重复签到记录，提示已签到", "", "未测", "", "", "", ""],
    ["SG-E2E-034", "签到消息", "P1", "通知", "B 已签到", "B 打开消息页", "可看到签到成功/活动提醒类消息，或有合理状态反馈", "", "未测", "", "", "", ""],
    ["SG-E2E-035", "分享活动", "P1", "增长", "活动详情可用", "点击分享按钮", "微信转发路径包含活动 id，分享卡片标题正确", "", "未测", "", "", "", ""],
    ["SG-E2E-036", "分享落地", "P1", "增长", "另一微信号收到分享", "从分享卡片进入", "直接打开对应活动详情页", "", "未测", "", "", "", ""],
    ["SG-E2E-037", "海报页", "P1", "增长", "活动详情可用", "点击生成海报", "进入海报页，展示封面、标题、时间地点、路径/占位码", "", "未测", "", "", "", ""],
    ["SG-E2E-038", "保存海报", "P2", "能力", "真机测试", "点击保存海报", "授权成功则保存相册；失败则提示清晰，不影响转发", "", "未测", "", "", "", ""],
    ["SG-E2E-039", "用户中心", "P1", "功能", "A/B 已登录", "进入用户中心", "显示头像、昵称、活动、评价、订单等内容", "", "未测", "", "", "", ""],
    ["SG-E2E-040", "编辑资料", "P1", "功能", "用户已登录", "修改昵称、简介、MBTI、头像后保存", "保存成功，重新进入用户中心仍显示新资料", "", "未测", "", "", "", ""],
    ["SG-E2E-041", "我的活动", "P0", "主流程", "A 创建、B 参加", "分别进入我的活动页", "A 在主办中看到活动，B 在参加/申请中看到记录", "", "未测", "", "", "", ""],
    ["SG-E2E-042", "消息已读", "P2", "通知", "消息页有未读消息", "点击消息或全部已读", "未读状态更新，刷新后仍保持", "", "未测", "", "", "", ""],
    ["SG-E2E-043", "举报活动", "P1", "治理", "B 在详情页", "点击举报该活动，提交原因", "举报成功，生成举报记录", "", "未测", "", "", "", ""],
    ["SG-E2E-044", "运营台入口", "P1", "运营", "A 有 admin/operator 角色", "A 进入用户中心", "显示运营入口，可进入运营台", "", "未测", "", "", "", ""],
    ["SG-E2E-045", "运营看板", "P1", "运营", "已有活动/报名/订单/签到数据", "A 打开运营台", "看板显示活动数、报名数、订单、签到率等", "", "未测", "", "", "", ""],
    ["SG-E2E-046", "举报处理", "P1", "治理", "有举报记录", "A 进入举报页，处理为已解决/驳回", "举报状态更新，相关用户收到系统通知", "", "未测", "", "", "", ""],
    ["SG-E2E-047", "活动下架", "P1", "治理", "A 有运营权限", "运营台对活动执行下架", "活动状态变为下架，详情/列表有对应提示或不可报名", "", "未测", "", "", "", ""],
    ["SG-E2E-048", "活动恢复", "P2", "治理", "活动已下架", "运营台恢复展示", "活动重新可见，状态一致", "", "未测", "", "", "", ""],
]

negative_rows = [
    ["SG-NEG-001", "未登录报名", "P0", "异常", "清缓存或退出登录", "进入详情点击报名", "跳转登录页，登录后返回报名流程", "", "未测", "", "", "", ""],
    ["SG-NEG-002", "未登录创建", "P0", "异常", "清缓存或退出登录", "点击底部 +", "跳转登录页，登录后进入创建页", "", "未测", "", "", "", ""],
    ["SG-NEG-003", "云函数失败", "P0", "异常", "临时断网或云函数异常", "打开核心页或提交操作", "页面不白屏，提示云端调用失败", "", "未测", "", "", "", ""],
    ["SG-NEG-004", "空活动列表", "P2", "边界", "清空测试活动数据", "进入首页/发现页", "显示空状态，不报错", "", "未测", "", "", "", ""],
    ["SG-NEG-005", "活动满员", "P1", "边界", "活动人数达到上限", "新用户尝试报名", "提示名额已满，不新增申请", "", "未测", "", "", "", ""],
    ["SG-NEG-006", "活动已结束", "P1", "边界", "活动状态为已结束", "用户尝试报名", "不允许报名，显示状态说明", "", "未测", "", "", "", ""],
    ["SG-NEG-007", "审核拒绝", "P1", "异常", "B 已提交申请", "A 拒绝 B 申请并填写原因", "B 状态为拒绝，不能进入支付/签到，消息展示原因", "", "未测", "", "", "", ""],
    ["SG-NEG-008", "订单关闭", "P1", "异常", "B 有待支付订单", "B 关闭订单", "状态为 closed，不能继续签到，详情只读", "", "未测", "", "", "", ""],
    ["SG-NEG-009", "模拟退款", "P1", "异常", "B 有已支付订单", "执行退款状态流转", "用户中心、参与者中心、订单详情均显示 refunded", "", "未测", "", "", "", ""],
    ["SG-NEG-010", "无相册权限", "P2", "权限", "真机拒绝保存相册授权", "点击保存海报", "显示授权失败提示，不影响分享/复制路径", "", "未测", "", "", "", ""],
    ["SG-NEG-011", "位置权限失败", "P2", "权限", "拒绝位置授权", "创建/详情使用地图能力", "有提示，文本地点仍可用", "", "未测", "", "", "", ""],
    ["SG-NEG-012", "扫码权限失败", "P2", "权限", "真机拒绝扫码或模拟器点击扫码", "A 点击扫码核销", "保留手动核销 fallback", "", "未测", "", "", "", ""],
]

ui_rows = [
    ["SG-UI-001", "iPhone 小屏", "P1", "兼容", "微信开发者工具/真机小屏", "预览首页、详情、创建、订单、海报", "无横向滚动、按钮不遮挡", "", "未测", "", "", "", ""],
    ["SG-UI-002", "Android 普通屏", "P1", "兼容", "Android 真机", "完整体验核心流程", "布局稳定，底部 Dock 不挡 CTA", "", "未测", "", "", "", ""],
    ["SG-UI-003", "大屏手机", "P2", "兼容", "大屏设备", "查看首页、发现、管理页", "内容宽度自然，卡片不拉伸失控", "", "未测", "", "", "", ""],
    ["SG-UI-004", "图片加载", "P1", "资源", "合法域名已配置", "查看封面、头像、海报", "云存储图片正常显示，外链失败有兜底", "", "未测", "", "", "", ""],
    ["SG-UI-005", "Toast 反馈", "P2", "体验", "核心按钮可用", "点击提交、审核、支付、签到、保存", "均有成功/失败反馈", "", "未测", "", "", "", ""],
    ["SG-UI-006", "返回链路", "P2", "体验", "各页面可进入", "使用页面返回/系统返回", "不出现死路或空白页", "", "未测", "", "", "", ""],
]

static_headers = ["检查ID", "检查项", "优先级", "检查方法", "预期结果", "实际结果", "状态", "缺陷ID", "负责人", "备注"]
static_rows = [
    ["SG-CHK-001", "Smoke check", "P0", "运行 node scripts\\surego-smoke-check.mjs", "通过", "", "未测", "", "", ""],
    ["SG-CHK-002", "Release check", "P0", "运行 node scripts\\surego-release-check.mjs", "通过", "", "未测", "", "", ""],
    ["SG-CHK-003", "Cloud integration static check", "P0", "运行 node scripts\\surego-cloud-integration-check.mjs", "通过", "", "未测", "", "", ""],
    ["SG-CHK-004", "微信开发者工具编译", "P0", "发行后导入 build/mp-weixin 并编译", "无编译错误", "", "未测", "", "", ""],
    ["SG-CHK-005", "合法域名", "P0", "真机/体验版观察控制台和页面资源", "request/upload/download 无域名报错", "", "未测", "", "", ""],
    ["SG-CHK-006", "控制台错误", "P1", "运行核心流程并查看 Console", "无 P0/P1 JS 报错", "", "未测", "", "", ""],
]

pre_headers = ["项目", "要求", "验证方法", "实际结果", "状态", "备注"]
pre_rows = [
    ["体验版", "已通过 HBuilderX 发行并上传体验版", "微信公众平台版本管理查看", "", "未测", ""],
    ["测试成员", "微信号 A/B 已加入体验成员", "成员管理查看并扫码验证", "", "未测", ""],
    ["AppID", "manifest mp-weixin.appid 已配置", "HBuilderX manifest 或项目配置查看", "", "未测", ""],
    ["uniCloud", "服务空间绑定正确", "HBuilderX uniCloud 目录显示目标空间", "", "未测", ""],
    ["云函数", "uni-id-co、uni-config-center、surego-* 已上传", "HBuilderX 云函数列表查看", "", "未测", ""],
    ["数据库", "SureGo schema 已上传，权限规则生效", "uniCloud 控制台查看", "", "未测", ""],
    ["合法域名", "request/upload/download 域名已配置", "微信公众平台服务器域名查看", "", "未测", ""],
    ["支付", "真实支付不接入，只测试运营订单确认", "支付页文案和订单状态检查", "", "确认", ""],
    ["运营权限", "测试运营台时 A 需有 admin/operator 角色", "uni-id 用户角色或 mock 权限查看", "", "未测", ""],
]

account_headers = ["账号", "角色", "用途", "微信昵称/备注", "测试设备", "登录结果", "UID", "备注"]
account_rows = [
    ["微信号 A", "发起人 / 局长 / 可选运营", "创建活动、审核报名、核销签到、运营处理", "", "", "", "", ""],
    ["微信号 B", "普通用户 / 参与者", "浏览、报名、支付确认、签到、查看消息", "", "", "", "", ""],
]

criteria_headers = ["等级", "通过标准", "判定结果", "备注"]
criteria_rows = [
    ["可进入体验测试", "P0 全部通过，P1 无阻塞或有明确规避方案", "", ""],
    ["可小范围试运营", "P0 全部通过，P1 通过率 >= 95%，真实云端数据跨账号一致", "", ""],
    ["暂停试运营", "任一 P0 失败，或登录/创建/报名/审核/支付/签到主链路断裂", "", ""],
]


def create_summary(wb):
    ws = wb.active
    ws.title = "00_说明与汇总"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "SureGo 体验版完整闭环测试用例"
    ws["A1"].font = Font(size=18, bold=True, color=THEME["navy"])
    ws["A2"] = "适用范围：体验版 / uniCloud 云端试运营模式 / 不接真实微信支付"
    ws["A3"] = "填写说明：执行测试后在各用例 Sheet 填写实际结果、状态、缺陷ID、负责人和备注。"

    ws["A5"] = "统计项"
    ws["B5"] = "数量"
    ws["A5"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["B5"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["A5"].font = ws["B5"].font = Font(color="FFFFFF", bold=True)
    stats = [
        ["E2E 主链路用例数", '=COUNTA(\'03_E2E闭环\'!A2:A200)'],
        ["异常边界用例数", '=COUNTA(\'04_异常边界\'!A2:A200)'],
        ["UI兼容用例数", '=COUNTA(\'05_UI兼容\'!A2:A200)'],
        ["静态检查项数", '=COUNTA(\'06_静态检查\'!A2:A200)'],
        ["P0 用例数", '=COUNTIF(\'03_E2E闭环\'!C:C,"P0")+COUNTIF(\'04_异常边界\'!C:C,"P0")+COUNTIF(\'05_UI兼容\'!C:C,"P0")+COUNTIF(\'06_静态检查\'!C:C,"P0")'],
        ["失败/阻塞数", '=COUNTIF(\'03_E2E闭环\'!I:I,"失败")+COUNTIF(\'03_E2E闭环\'!I:I,"阻塞")+COUNTIF(\'04_异常边界\'!I:I,"失败")+COUNTIF(\'04_异常边界\'!I:I,"阻塞")+COUNTIF(\'05_UI兼容\'!I:I,"失败")+COUNTIF(\'05_UI兼容\'!I:I,"阻塞")+COUNTIF(\'06_静态检查\'!G:G,"失败")+COUNTIF(\'06_静态检查\'!G:G,"阻塞")'],
    ]
    for row in stats:
        ws.append(row)

    ws["D5"] = "试运营建议"
    ws["E5"] = "判定"
    ws["D5"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["E5"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["D5"].font = ws["E5"].font = Font(color="FFFFFF", bold=True)
    ws["D6"] = "若失败/阻塞数为 0 且 P0 全通过，可进入小范围体验测试。"
    ws["E6"] = '=IF(B11=0,"可继续体验版测试","需先处理失败/阻塞")'

    for row in ws.iter_rows(min_row=1, max_row=20, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    set_widths(ws, {"A": 26, "B": 18, "C": 4, "D": 48, "E": 24, "F": 4})


def add_case_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    append_rows(ws, e2e_headers, rows)
    set_widths(ws, {
        "A": 15, "B": 18, "C": 8, "D": 12, "E": 28, "F": 42, "G": 42,
        "H": 30, "I": 10, "J": 14, "K": 12, "L": 16, "M": 24,
    })
    style_sheet(ws, title.replace("_", "")[:20])
    add_validations(ws)
    return ws


def main():
    wb = Workbook()
    create_summary(wb)

    ws = wb.create_sheet("01_前置检查")
    append_rows(ws, pre_headers, pre_rows)
    set_widths(ws, {"A": 18, "B": 42, "C": 34, "D": 24, "E": 10, "F": 28})
    style_sheet(ws, "PreflightTable")
    dv = DataValidation(type="list", formula1='"未测,通过,失败,阻塞,跳过,确认"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{ws.max_row}")

    ws = wb.create_sheet("02_测试账号")
    append_rows(ws, account_headers, account_rows)
    set_widths(ws, {"A": 14, "B": 26, "C": 46, "D": 20, "E": 18, "F": 18, "G": 28, "H": 28})
    style_sheet(ws, "AccountsTable")

    add_case_sheet(wb, "03_E2E闭环", e2e_rows)
    add_case_sheet(wb, "04_异常边界", negative_rows)
    add_case_sheet(wb, "05_UI兼容", ui_rows)

    ws = wb.create_sheet("06_静态检查")
    append_rows(ws, static_headers, static_rows)
    set_widths(ws, {"A": 15, "B": 22, "C": 8, "D": 44, "E": 34, "F": 28, "G": 10, "H": 14, "I": 12, "J": 24})
    style_sheet(ws, "StaticChecksTable")
    dv_status = DataValidation(type="list", formula1='"未测,通过,失败,阻塞,跳过"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_priority)
    dv_status.add(f"G2:G{ws.max_row}")
    dv_priority.add(f"C2:C{ws.max_row}")

    ws = wb.create_sheet("07_通过标准")
    append_rows(ws, criteria_headers, criteria_rows)
    set_widths(ws, {"A": 18, "B": 72, "C": 20, "D": 34})
    style_sheet(ws, "CriteriaTable")

    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 42 if row > 1 else 28
        sheet.freeze_panes = "A2" if sheet.title != "00_说明与汇总" else None

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
